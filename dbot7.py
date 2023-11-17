import openai
import os
from pathlib import Path
from tkinter import filedialog, messagebox
from openpyxl.styles import Alignment, Border, Font, PatternFill
import shutil
import openpyxl


openai.api_key = os.getenv("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")
def get_website_links_from_workbook(workbook_path):
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    if "Qualitative Review" not in workbook.sheetnames:
        return [], None, False
    sheet = workbook["Qualitative Review"]
    for row_idx, row in enumerate(sheet.iter_rows(min_col=1, max_col=sheet.max_column), 1):
        if row[0].value == '#':
            break_row = row_idx
            break
    website_col_idx = None
    for col_idx, cell in enumerate(sheet[break_row], 1):
        if cell.value == "Website Address":
            website_col_idx = col_idx
            break
    if not website_col_idx:
        return [], None, False
    links = [cell[0].value for cell in
             sheet.iter_rows(min_row=break_row + 1, min_col=website_col_idx, max_col=website_col_idx) if cell[0].value]
    return links, workbook, True


def process_search_matrix(sheet):
    for row_idx, row in enumerate(sheet.iter_rows(min_col=1, max_col=1), 1):
        if row[0].value == '#':
            break_row = row_idx
            break
    allowed_columns = ['#', 'Company Name', 'Accept/Reject', 'Comments']
    accept_reject_col = None
    comments_col = None
    for col_idx, cell in enumerate(sheet[break_row], 1):
        if cell.value == 'Accept/Reject':
            accept_reject_col = col_idx
        elif cell.value == 'Comments':
            comments_col = col_idx
    cols_to_delete = []
    for col_idx, col in enumerate(sheet.iter_cols(min_row=break_row - 1, max_row=break_row), 1):
        if accept_reject_col <= col_idx <= comments_col and col[0].value != 0:
            continue
        elif col[1].value not in allowed_columns:
            cols_to_delete.append(col_idx)
    for col_idx in sorted(cols_to_delete, reverse=True):
        sheet.delete_cols(col_idx)
    sheet.title = 'Search Matrix'


def process_business_descriptions(sheet):
    for row_idx, row in enumerate(sheet.iter_rows(min_col=1, max_col=1), 1):
        if row[0].value == '#':
            break_row = row_idx
            break
    allowed_columns = ['#', 'Company Name', 'Full Overview']
    cols_to_delete = []
    for col_idx, col in enumerate(sheet.iter_cols(min_row=break_row - 1, max_row=break_row), 1):
        if col[1].value not in allowed_columns:
            cols_to_delete.append(col_idx)
    for col_idx in sorted(cols_to_delete, reverse=True):
        sheet.delete_cols(col_idx)
    last_number_row = break_row
    for row in sheet.iter_rows(min_row=break_row + 1, min_col=1, max_col=1):
        cell = row[0]
        if isinstance(cell.value, int) or (isinstance(cell.value, str) and cell.value.isdigit()):
            last_number_row = cell.row
    if sheet.max_row > last_number_row:
        sheet.delete_rows(last_number_row + 1, sheet.max_row - last_number_row)
    for col_idx, col in enumerate(sheet.iter_cols(min_row=break_row, max_row=break_row), 1):
        if col[0].value == 'Company Name':
            for cell in sheet.iter_rows(min_col=col_idx, max_col=col_idx, min_row=break_row + 1):
                cell[0].alignment = Alignment(vertical='center')
        elif col[0].value == 'Full Overview':
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = 500 / 7
    sheet.title = 'Business Descriptions'


def process_ratio_report():
    sample_path = filedialog.askopenfilename(title="Select Sample Workbook",
                                             filetypes=[("Excel files", "*.xlsx;*.xls")])
    if not sample_path:
        return
    target_paths = filedialog.askopenfilenames(title="Select Target Workbooks",
                                               filetypes=[("Excel files", "*.xlsx;*.xls")])
    if not target_paths:
        return
    sample_wb = openpyxl.load_workbook(sample_path)
    sample_sheet = sample_wb.active
    for target_path in target_paths:
        target_wb = openpyxl.load_workbook(target_path)
        for target_sheet in target_wb:
            for row_sample, row_target in zip(sample_sheet.iter_rows(), target_sheet.iter_rows()):
                for cell_sample, cell_target in zip(row_sample, row_target):
                    if cell_sample.has_style:
                        font = Font(name=cell_sample.font.name, size=cell_sample.font.size,
                                    bold=cell_sample.font.bold, italic=cell_sample.font.italic,
                                    vertAlign=cell_sample.font.vertAlign, underline=cell_sample.font.underline,
                                    strike=cell_sample.font.strike, color=cell_sample.font.color)
                        alignment = Alignment(horizontal=cell_sample.alignment.horizontal,
                                              vertical=cell_sample.alignment.vertical,
                                              text_rotation=cell_sample.alignment.text_rotation,
                                              wrap_text=cell_sample.alignment.wrap_text,
                                              shrink_to_fit=cell_sample.alignment.shrink_to_fit,
                                              indent=cell_sample.alignment.indent)
                        border = Border(left=cell_sample.border.left, right=cell_sample.border.right,
                                        top=cell_sample.border.top, bottom=cell_sample.border.bottom)
                        fill = PatternFill(fill_type=cell_sample.fill.fill_type,
                                           start_color=cell_sample.fill.start_color,
                                           end_color=cell_sample.fill.end_color)
                        cell_target.font = font
                        cell_target.border = border
                        cell_target.fill = fill
                        cell_target.alignment = alignment
                        cell_target.number_format = cell_sample.number_format
            for col_sample, col_target in zip(sample_sheet.iter_cols(), target_sheet.iter_cols()):
                col_letter_sample = openpyxl.utils.get_column_letter(col_sample[0].column)
                col_letter_target = openpyxl.utils.get_column_letter(col_target[0].column)
                if sample_sheet.column_dimensions[col_letter_sample].hidden:
                    target_sheet.column_dimensions[col_letter_target].hidden = True
            for row_sample, row_target in zip(sample_sheet.iter_rows(), target_sheet.iter_rows()):
                if sample_sheet.row_dimensions[row_sample[0].row].hidden:
                    target_sheet.row_dimensions[row_target[0].row].hidden = True
        target_wb.save(target_path)


# def overview_review()
# def dbhoovers_check()
# def website_review()
# def web_vs_overview()
# reasoning_generator()

def perform_qualitative_review(file_paths):
    # Iterate through each company in the workbook
    # Perform the specified actions based on the provided criteria
    for file_path in file_paths:
        website_links, _, links_found = get_website_links_from_workbook(file_path)
        if not links_found:
            messagebox.showinfo("No Links Found",
                                f"The program couldn't find any website links in the workbook: {file_path}.")
            continue
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        if "Qualitative Review" not in workbook.sheetnames:
            continue
        sheet = workbook["Qualitative Review"]
        for row_idx, row in enumerate(sheet.iter_rows(min_col=1, max_col=1), 1):
            if row[0].value == '#':
                break_row = row_idx
                break
        website_col_idx = None
        description_col_idx = None
        status_col_idx = None
        for col_idx, cell in enumerate(sheet[break_row], 1):
            if cell.value == "Website Address":
                website_col_idx = col_idx
            elif cell.value == "Description from website":
                description_col_idx = col_idx
            elif cell.value == "status":
                status_col_idx = col_idx
        if not website_col_idx or not description_col_idx:
            continue
        for row_idx, link in enumerate(website_links, start=1):
            if sheet.cell(row=break_row + row_idx, column=status_col_idx).value != "Reject":
                prompt = f"""
                Using GPT-4, analyze the business services provided by the company at the following link: {link}. 
                Please provide a summary of the services in bullet-point format, accompanied by the sources 
                from where each insight was determined.

                The summary should be structured as follows:

                [Company Name] is a company that provides a variety of [industry/type]-related services. 
                Their offerings include:

                - [Service 1]: (Source: [Link to the specific page or section where this information was found])
                - [Service 2]: (Source: [Link to the specific page or section where this information was found])
                - [Service 3]: (Source: [Link to the specific page or section where this information was found])
                - [Service 4]: (Source: [Link to the specific page or section where this information was found])
                """

                try:
                    sheet.cell(row=break_row + row_idx, column=description_col_idx).value = openai.Completion.create(
                        prompt, max_tokens=150)
                except Exception:
                    sheet.cell(row=break_row + row_idx, column=description_col_idx).value = "Unable to process website"
        workbook.save(file_path)
        """
        company_name = company_row["Company"]  # Assuming 'Company' is a column name
        full_overview = company_row["Full Overview"]  # Assuming 'Full Overview' is a column name

        # 1) Compare Full Overview with positive and negative text
        similarity = compare_texts(full_overview, positive_text, negative_text)

        if similarity > threshold:
            company_row["Detailed Review"] = "X"  # Mark 'X' for detailed review
        else:
            company_row["Status"] = "Rejected"  # Mark 'Rejected'

        # 2) Check DB Hoovers independence status
        if company_row["Detailed Review"] == "X":
            independence_status = check_db_hoovers_independence(company_name)

            # Update status accordingly
            if independence_status == "Subsidiary":
                company_row["Status"] = "Rejected"


        # 3) Review websites and highlight differences
        # similarity_to_party_type = compare_to_party_type(website_description, selected_value_party)

                    website_similarity = compare_website_to_full_overview(company_row["Website"], full_overview)
                    if website_similarity < website_threshold:
                        highlight_website_difference(company_row["Full Overview"])
        """


def process_request(tasks_list):
    if tasks_list.Contains() is None:
        messagebox.showwarning("No Option Selected", "Please select an option before processing.")
        return
    else:
        if tasks_list.contains("review"):
            file_paths = filedialog.askopenfilenames(title="Select Workbooks for Qualitative Review",
                                                     filetypes=[("Excel files", "*.xlsx;*.xls")])
            if file_paths:
                perform_qualitative_review(file_paths)

        if tasks_list.contains("appendices"):
            process_ratio_report()
            file_paths = filedialog.askopenfilenames(title="Select Workbooks for other Appendices",
                                                     filetypes=[("Excel files", "*.xlsx;*.xls")])
            if file_paths:
                for file_path in file_paths:
                    workbook = openpyxl.load_workbook(file_path, data_only=True)
                    keep_sheets = []
                    if "Draft Results" in workbook.sheetnames:
                        sheet = workbook["Draft Results"]
                        process_business_descriptions(sheet)
                        keep_sheets.append("Business Descriptions")
                    if "Qualitative Review" in workbook.sheetnames:
                        sheet = workbook["Qualitative Review"]
                        process_search_matrix(sheet)
                        keep_sheets.append("Search Matrix")
                    for sheet_name in workbook.sheetnames:
                        if sheet_name not in keep_sheets:
                            del workbook[sheet_name]
                    workbook.save(file_path)
                    output_folder = Path.home() / "Downloads" / "dbot output"
                    output_folder.mkdir(parents=True, exist_ok=True)
                    output_file_path = output_folder / Path(file_path).name
                    shutil.move(file_path, output_file_path)


